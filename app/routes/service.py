import logging
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from app.extensions import db
from app.models.user import User
from app.models.calculation import CalculationHistory
from app.decorators import service_required
from app.utils.excel_import import ExcelImporter
from app.utils.exceptions import ValidationError, ResourceNotFoundError, BusinessError, PermissionDeniedError
from datetime import datetime
import os
import tempfile
from openpyxl import load_workbook

service_bp = Blueprint('service', __name__)
logger = logging.getLogger(__name__)

@service_bp.route('/customers', methods=['GET'])
@login_required
@service_required
def get_customers():
    """获取客户列表"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        query = request.args.get('query', '')
        status = request.args.get('status')
        
        # 验证分页参数
        if page < 1:
            raise ValidationError('页码必须大于0')
        if per_page < 1:
            raise ValidationError('每页数量必须大于0')
            
        # 构建查询
        customers_query = User.query.filter_by(role='customer')
        
        # 搜索过滤
        if query:
            customers_query = customers_query.filter(
                db.or_(
                    User.username.ilike(f'%{query}%'),
                    User.email.ilike(f'%{query}%')
                )
            )
            
        # 状态过滤
        if status:
            if status not in ['active', 'inactive']:
                raise ValidationError('无效的状态值')
            customers_query = customers_query.filter_by(is_active=(status == 'active'))
            
        # 分页
        pagination = customers_query.order_by(User.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        customers = [customer.to_dict() for customer in pagination.items]
        logger.debug(f"获取客户列表成功，数量: {len(customers)}")
        
        return jsonify({
            'customers': customers,
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        })
        
    except ValidationError as e:
        logger.warning(f"获取客户列表验证错误: {str(e)}")
        return jsonify({'message': str(e)}), 400
    except Exception as e:
        logger.error(f"获取客户列表失败: {str(e)}", exc_info=True)
        return jsonify({'message': '获取客户列表失败'}), 500

@service_bp.route('/customers/<int:id>', methods=['GET'])
@login_required
@service_required
def get_customer(id):
    """获取客户详情"""
    try:
        customer = User.query.get(id)
        if not customer:
            raise ResourceNotFoundError(f'客户不存在: {id}')
            
        if customer.role != 'customer':
            raise BusinessError('无效的客户ID')
            
        logger.debug(f"获取客户详情成功: {id}")
        return jsonify(customer.to_dict())
        
    except ResourceNotFoundError as e:
        logger.warning(str(e))
        return jsonify({'message': str(e)}), 404
    except BusinessError as e:
        logger.warning(f"获取客户详情业务错误: {str(e)}")
        return jsonify({'message': str(e)}), 400
    except Exception as e:
        logger.error(f"获取客户详情失败: {str(e)}", exc_info=True)
        return jsonify({'message': '获取客户详情失败'}), 500

@service_bp.route('/customers/<int:id>/history', methods=['GET'])
@login_required
@service_required
def get_customer_history(id):
    """获取客户计算历史"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        
        # 验证分页参数
        if page < 1:
            raise ValidationError('页码必须大于0')
        if per_page < 1:
            raise ValidationError('每页数量必须大于0')
            
        # 验证客户
        customer = User.query.get(id)
        if not customer:
            raise ResourceNotFoundError(f'客户不存在: {id}')
            
        if customer.role != 'customer':
            raise BusinessError('无效的客户ID')
            
        # 日期范围过滤
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 构建查询
        query = CalculationHistory.query.filter_by(user_id=id)
        
        if start_date:
            try:
                start_date = datetime.fromisoformat(start_date)
                query = query.filter(CalculationHistory.created_at >= start_date)
            except ValueError:
                raise ValidationError('无效的开始日期格式，请使用ISO格式(YYYY-MM-DD)')
                
        if end_date:
            try:
                end_date = datetime.fromisoformat(end_date)
                query = query.filter(CalculationHistory.created_at <= end_date)
            except ValueError:
                raise ValidationError('无效的结束日期格式，请使用ISO格式(YYYY-MM-DD)')
                
        # 分页
        pagination = query.order_by(CalculationHistory.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        history = []
        for record in pagination.items:
            history_dict = record.to_dict()
            history_dict['product'] = record.product.to_dict() if record.product else None
            history.append(history_dict)
            
        logger.debug(f"获取客户计算历史成功，客户ID: {id}, 数量: {len(history)}")
        return jsonify({
            'customer': customer.to_dict(),
            'history': history,
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        })
        
    except ValidationError as e:
        logger.warning(f"获取客户计算历史验证错误: {str(e)}")
        return jsonify({'message': str(e)}), 400
    except ResourceNotFoundError as e:
        logger.warning(str(e))
        return jsonify({'message': str(e)}), 404
    except BusinessError as e:
        logger.warning(f"获取客户计算历史业务错误: {str(e)}")
        return jsonify({'message': str(e)}), 400
    except Exception as e:
        logger.error(f"获取客户计算历史失败: {str(e)}", exc_info=True)
        return jsonify({'message': '获取客户计算历史失败'}), 500

@service_bp.route('/customers/<int:id>/history/export', methods=['GET'])
@login_required
@service_required
def export_customer_history(id):
    """导出客户计算历史"""
    try:
        # 验证客户
        customer = User.query.get(id)
        if not customer:
            raise ResourceNotFoundError(f'客户不存在: {id}')
            
        if customer.role != 'customer':
            raise BusinessError('无效的客户ID')
            
        # 日期范围过滤
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 构建查询
        query = CalculationHistory.query.filter_by(user_id=id)
        
        if start_date:
            try:
                start_date = datetime.fromisoformat(start_date)
                query = query.filter(CalculationHistory.created_at >= start_date)
            except ValueError:
                raise ValidationError('无效的开始日期格式，请使用ISO格式(YYYY-MM-DD)')
                
        if end_date:
            try:
                end_date = datetime.fromisoformat(end_date)
                query = query.filter(CalculationHistory.created_at <= end_date)
            except ValueError:
                raise ValidationError('无效的结束日期格式，请使用ISO格式(YYYY-MM-DD)')
                
        history = query.order_by(CalculationHistory.created_at.desc()).all()
        if not history:
            raise BusinessError('没有可导出的数据')
            
        # 创建临时文件
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, f'customer_{id}_history.xlsx')
        
        try:
            # 导出Excel
            with ExcelImporter.create_workbook(temp_path) as workbook:
                worksheet = workbook.active
                worksheet.title = '计算历史'
                
                # 写入表头
                headers = [
                    'ID', '产品', '长度', '宽度', '高度', '重量', '分区',
                    '是否住宅', '基础费用', '操作费', '超大费用',
                    '住宅费用', '偏远费用', '总费用', '计算时间'
                ]
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col, value=header)
                    
                # 写入数据
                for row, record in enumerate(history, 2):
                    product = record.product
                    worksheet.cell(row=row, column=1, value=record.id)
                    worksheet.cell(row=row, column=2, value=product.name if product else None)
                    worksheet.cell(row=row, column=3, value=record.length)
                    worksheet.cell(row=row, column=4, value=record.width)
                    worksheet.cell(row=row, column=5, value=record.height)
                    worksheet.cell(row=row, column=6, value=record.weight)
                    worksheet.cell(row=row, column=7, value=record.zone)
                    worksheet.cell(row=row, column=8, value='是' if record.is_residential else '否')
                    worksheet.cell(row=row, column=9, value=record.base_fee)
                    worksheet.cell(row=row, column=10, value=record.handling_fee)
                    worksheet.cell(row=row, column=11, value=record.oversize_fee)
                    worksheet.cell(row=row, column=12, value=record.residential_fee)
                    worksheet.cell(row=row, column=13, value=record.remote_area_fee)
                    worksheet.cell(row=row, column=14, value=record.total_fee)
                    worksheet.cell(row=row, column=15, value=record.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                    
            logger.info(f"导出客户计算历史成功，客户ID: {id}, 数量: {len(history)}")
            return send_file(
                temp_path,
                as_attachment=True,
                download_name=f'customer_{id}_history.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        finally:
            # 清理临时文件
            try:
                os.remove(temp_path)
                os.rmdir(temp_dir)
            except Exception as e:
                logger.warning(f"清理临时文件失败: {str(e)}")
                
    except ValidationError as e:
        logger.warning(f"导出客户计算历史验证错误: {str(e)}")
        return jsonify({'message': str(e)}), 400
    except ResourceNotFoundError as e:
        logger.warning(str(e))
        return jsonify({'message': str(e)}), 404
    except BusinessError as e:
        logger.warning(f"导出客户计算历史业务错误: {str(e)}")
        return jsonify({'message': str(e)}), 400
    except Exception as e:
        logger.error(f"导出客户计算历史失败: {str(e)}", exc_info=True)
        return jsonify({'message': '导出客户计算历史失败'}), 500

# 错误处理器
@service_bp.errorhandler(ValidationError)
def handle_validation_error(e):
    """处理验证错误"""
    logger.warning(f"验证错误: {str(e)}")
    return jsonify({'message': str(e)}), 400

@service_bp.errorhandler(ResourceNotFoundError)
def handle_not_found_error(e):
    """处理资源不存在错误"""
    logger.warning(str(e))
    return jsonify({'message': str(e)}), 404

@service_bp.errorhandler(BusinessError)
def handle_business_error(e):
    """处理业务错误"""
    logger.warning(f"业务错误: {str(e)}")
    return jsonify({'message': str(e)}), 400

@service_bp.errorhandler(PermissionDeniedError)
def handle_permission_denied(e):
    """处理权限不足错误"""
    logger.warning(f"权限不足: {str(e)}")
    return jsonify({'message': str(e)}), 403

@service_bp.errorhandler(Exception)
def handle_unexpected_error(e):
    """处理未预期的错误"""
    logger.error(f"未预期的错误: {str(e)}", exc_info=True)
    return jsonify({'message': '服务器内部错误'}), 500 